# -*- coding: utf-8 -*-
"""
📊 Excel 분석 워크북 생성 모듈.

매 실행마다 yfinance에서 최근 2년 일별 종가를 받아서:
- Prices        : 일별 종가 (값)
- Returns       : 일별 수익률 (Excel 공식)
- Analysis      : CFA 지표 대시보드 (Excel 공식) — 수익률/변동성/샤프/이동평균 등
- Correlation   : 자산 간 상관계수 (Excel 공식)
- 검증(Python)   : 같은 지표를 pandas로 계산한 값 (공식 결과와 대조용)

핵심 함수: build_analysis_workbook(tickers, path=None, period="2y") -> (경로, 에러 or None)

설계 원칙:
- 매번 Prices를 새로 채움 → 누적 상태 버그 없음, 항상 완전한 시계열
- 공식은 넉넉한 범위(2~100000행)를 참조 → 데이터가 늘어도 자동 반영
- '검증' 시트로 공식 vs 파이썬 계산을 눈으로 대조 가능
- 사용자가 직접 만든 다른 시트(예: 메모)는 건드리지 않음
"""

import os
import warnings
warnings.filterwarnings("ignore")
from datetime import datetime

import numpy as np
import pandas as pd
import yfinance as yf
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DEFAULT_PATH = os.path.join(BASE_DIR, "MorningBrief_Analysis.xlsx")

# 자동으로 다시 쓰는 시트들 (이 이름들은 매번 새로 만듦, 나머지는 보존)
# 'Alerts'(애드인이 만든 수식 시트)는 여기 없음 → 건드리지 않고 보존됨
MANAGED_SHEETS = ["애널리스트분석", "자동알림", "Analysis", "Prices", "Returns", "Correlation", "검증(Python)"]

TRADING_DAYS = 252          # 연환산에 쓰는 거래일 수
RISK_FREE = 0.03            # 무위험수익률(연) 기본값 = 3% (한은 기준금리 근처)

# ---- 스타일 ----
FONT = "Arial"
HDR_FILL = PatternFill("solid", fgColor="4F46E5")   # 남보라 헤더
HDR_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name=FONT, bold=True, size=14, color="1F2937")
INPUT_FONT = Font(name=FONT, bold=True, color="0000FF")   # 파랑 = 입력값(수정 가능)
INPUT_FILL = PatternFill("solid", fgColor="FFF3B0")       # 노랑 = 사용자 입력 셀
BASE_FONT = Font(name=FONT, size=10)
THIN = Side(style="thin", color="DDDDDD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _market_class(symbol: str) -> str:
    """자산을 시장별로 분류 (완료봉 판정·시차 처리에 사용)."""
    if symbol.endswith(".KS") or symbol.endswith(".KQ") or symbol == "^KS11":
        return "KR"
    if symbol.endswith("=X"):
        return "FX"
    if symbol.endswith("-USD"):
        return "CRYPTO"
    return "US"   # 미국 지수/주식/원자재선물


def _last_bar_complete(symbol, bar_date, now_utc) -> bool:
    """
    해당 자산의 마지막 봉이 '완료된 정규장'인지 판정.
    - KR: 오늘 봉은 한국 정규장 마감(15:30 KST=06:30 UTC) 이후에만 완료
    - US/FX/CRYPTO: 오늘(UTC) 날짜 봉은 진행중으로 간주 → 직전 완료봉만 사용
      (오후 4시 KST 실행 시 미국은 자동으로 '새벽 5시에 끝난 정규장' 종가가 됨)
    """
    today = now_utc.date()
    if _market_class(symbol) == "KR":
        if bar_date == today:
            return now_utc >= (pd.Timestamp(today) + pd.Timedelta(hours=6, minutes=30))
        return True
    return bar_date < today


def _download_prices(tickers, period="2y", include_today=False):
    """
    yfinance로 여러 종목 종가를 받아 하나의 DataFrame으로 정렬.
    반환: (DataFrame[dates x name], 실패한종목리스트)
    """
    symbols = [t["symbol"] for t in tickers]
    names = [t["name"] for t in tickers]
    sym2name = dict(zip(symbols, names))

    # auto_adjust=False → 배당·분할 조정 안 된 '실제 종가'(Naver/야후 표시값과 동일)
    raw = yf.download(symbols, period=period, interval="1d",
                      auto_adjust=False, progress=False, group_by="column")

    # 여러 종목이면 raw["Close"]가 DataFrame, 단일이면 Series
    if isinstance(raw.columns, pd.MultiIndex):
        close = raw["Close"]
    else:
        close = raw[["Close"]]
        close.columns = symbols

    # 원래 순서 유지 + 존재하는 심볼만
    have = [s for s in symbols if s in close.columns]
    failed = [sym2name[s] for s in symbols if s not in have]
    close = close[have]
    # 완전히 빈 컬럼도 실패 처리
    for s in list(close.columns):
        if close[s].dropna().empty:
            failed.append(sym2name[s])
            close = close.drop(columns=[s])

    # 🔧 진행 중인 장중 봉 제거 — 각 자산의 마지막 봉이 미완료면 NaN 처리
    #    (미국 장중값 잘못 기록되는 문제 해결. ffill이 직전 완료 종가로 채움)
    now_utc = pd.Timestamp(datetime.utcnow())
    for s in list(close.columns):
        col = close[s].dropna()
        if not col.empty and not _last_bar_complete(s, col.index[-1].date(), now_utc):
            close.loc[col.index[-1], s] = np.nan

    close = close.rename(columns=sym2name)
    # 시장마다 휴장일 달라서 앞으로 채우기(ffill)
    close = close.sort_index().ffill()
    # 주말(토/일) 행 제거 — 암호화폐만 거래된 날(증시는 ffill이라 가짜 0% 유발)
    close = close[close.index.dayofweek < 5]
    # 오후 실행(include_today=True): 당일 국장 종가까지 포함
    # 아침 실행(False): 진행 중인 오늘을 빼고 직전 완료 거래일까지만
    today = pd.Timestamp(datetime.now().date())
    if include_today:
        close = close[close.index <= today]
    else:
        close = close[close.index < today]
    close = close.dropna(how="all")
    return close, failed


def _style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _write_prices(ws, prices: pd.DataFrame):
    """Prices 시트: 날짜 + 종가(값)."""
    names = list(prices.columns)
    ws.cell(row=1, column=1, value="Date")
    for j, name in enumerate(names):
        ws.cell(row=1, column=2 + j, value=name)
    _style_header(ws, 1, len(names) + 1)

    for i, (dt, row) in enumerate(prices.iterrows()):
        r = 2 + i
        dcell = ws.cell(row=r, column=1, value=dt.to_pydatetime())
        dcell.number_format = "yyyy-mm-dd"
        dcell.font = BASE_FONT
        for j, name in enumerate(names):
            val = row[name]
            cell = ws.cell(row=r, column=2 + j,
                           value=(None if pd.isna(val) else float(val)))
            cell.number_format = "#,##0.00"
            cell.font = BASE_FONT
    ws.freeze_panes = "B2"
    ws.column_dimensions["A"].width = 12
    for j in range(len(names)):
        ws.column_dimensions[get_column_letter(2 + j)].width = 13


def _write_returns(ws, prices: pd.DataFrame):
    """Returns 시트: 일별 수익률 (Excel 공식으로)."""
    names = list(prices.columns)
    nrows = len(prices)
    ws.cell(row=1, column=1, value="Date")
    for j, name in enumerate(names):
        ws.cell(row=1, column=2 + j, value=name)
    _style_header(ws, 1, len(names) + 1)

    for i in range(nrows):
        r = 2 + i
        # 날짜는 Prices에서 그대로 참조
        ws.cell(row=r, column=1, value=f"=Prices!A{r}").number_format = "yyyy-mm-dd"
        ws.cell(row=r, column=1).font = BASE_FONT
        for j in range(len(names)):
            cl = get_column_letter(2 + j)
            if i == 0:
                cell = ws.cell(row=r, column=2 + j, value=None)  # 첫날은 이전값 없음
            else:
                f = f"=IFERROR(Prices!{cl}{r}/Prices!{cl}{r-1}-1,\"\")"
                cell = ws.cell(row=r, column=2 + j, value=f)
            cell.number_format = "0.00%"
            cell.font = BASE_FONT
    ws.freeze_panes = "B2"
    ws.column_dimensions["A"].width = 12
    for j in range(len(names)):
        ws.column_dimensions[get_column_letter(2 + j)].width = 11


def _write_analysis(ws, prices: pd.DataFrame):
    """Analysis 시트: CFA 지표 대시보드 (전부 Excel 공식)."""
    names = list(prices.columns)

    ws.cell(row=1, column=1, value="📊 CFA Analysis Dashboard").font = TITLE_FONT
    ws.cell(row=2, column=1, value="무위험수익률 rf (연):").font = Font(name=FONT, bold=True)
    rf_cell = ws.cell(row=2, column=2, value=RISK_FREE)
    rf_cell.font = INPUT_FONT
    rf_cell.fill = INPUT_FILL
    rf_cell.number_format = "0.0%"
    rf_cell.border = BORDER
    ws.cell(row=2, column=3,
            value="← 파란 셀은 수정 가능 (샤프비율 계산에 사용). 예: 한은 기준금리").font = Font(name=FONT, italic=True, color="888888", size=9)

    # 헤더
    headers = ["자산", "최신가", "1일%", "1주%", "1개월%", "3개월%", "기간수익%",
               "연환산수익%", "연환산변동성%", "샤프비율",
               "MA20", "MA60", "MA200", "現價vsMA200%", "52주고점", "고점대비%"]
    hrow = 4
    for c, h in enumerate(headers, start=1):
        ws.cell(row=hrow, column=c, value=h)
    _style_header(ws, hrow, len(headers))

    # 퍼센트로 표시할 컬럼 인덱스(1-base)
    pct_cols = {3, 4, 5, 6, 7, 8, 9, 14, 16}
    price_cols = {2, 11, 12, 13, 15}

    for i, name in enumerate(names):
        r = hrow + 1 + i
        cl = get_column_letter(2 + i)           # Prices/Returns에서 이 자산의 컬럼
        R = f"Prices!${cl}$2:${cl}$100000"      # 가격 범위
        RR = f"Returns!${cl}$2:${cl}$100000"    # 수익률 범위
        n = f"COUNT({R})"                        # 데이터 개수 = 마지막 인덱스
        last = f"INDEX({R},{n})"

        def ago(k):   # k거래일 전 가격
            return f"INDEX({R},{n}-{k})"
        def ma(k):    # 최근 k일 이동평균
            return f"AVERAGE(INDEX({R},{n}-{k-1}):INDEX({R},{n}))"

        ws.cell(row=r, column=1, value=name).font = Font(name=FONT, bold=True)
        formulas = {
            2:  f"=IFERROR({last},\"\")",                                  # 최신가
            3:  f"=IFERROR({last}/{ago(1)}-1,\"\")",                        # 1일
            4:  f"=IFERROR({last}/{ago(5)}-1,\"\")",                        # 1주
            5:  f"=IFERROR({last}/{ago(21)}-1,\"\")",                       # 1개월
            6:  f"=IFERROR({last}/{ago(63)}-1,\"\")",                       # 3개월
            7:  f"=IFERROR({last}/INDEX({R},1)-1,\"\")",                    # 기간 전체
            8:  f"=IFERROR(AVERAGE({RR})*{TRADING_DAYS},\"\")",            # 연환산 수익
            9:  f"=IFERROR(STDEV({RR})*SQRT({TRADING_DAYS}),\"\")",        # 연환산 변동성
            10: f"=IFERROR((AVERAGE({RR})*{TRADING_DAYS}-$B$2)/(STDEV({RR})*SQRT({TRADING_DAYS})),\"\")",  # 샤프
            11: f"=IFERROR({ma(20)},\"\")",
            12: f"=IFERROR({ma(60)},\"\")",
            13: f"=IFERROR({ma(200)},\"\")",
            14: f"=IFERROR({last}/{ma(200)}-1,\"\")",                       # 현재가 vs MA200
            15: f"=IFERROR(MAX(INDEX({R},MAX(1,{n}-251)):INDEX({R},{n})),\"\")",  # 52주 고점
            16: f"=IFERROR({last}/MAX(INDEX({R},MAX(1,{n}-251)):INDEX({R},{n}))-1,\"\")",  # 고점대비
        }
        for c, f in formulas.items():
            cell = ws.cell(row=r, column=c, value=f)
            cell.font = BASE_FONT
            if c in pct_cols:
                cell.number_format = "0.00%"
            elif c in price_cols:
                cell.number_format = "#,##0.00"
            else:
                cell.number_format = "0.00"

    ws.freeze_panes = "B5"
    ws.column_dimensions["A"].width = 14
    for c in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 12


def _write_correlation(ws, prices: pd.DataFrame):
    """Correlation 시트: 자산 간 일별수익률 상관계수 (CORREL 공식)."""
    names = list(prices.columns)
    ws.cell(row=1, column=1, value="자산 간 상관계수 (일별 수익률 기준)").font = TITLE_FONT
    hrow = 3
    ws.cell(row=hrow, column=1, value="")
    for j, name in enumerate(names):
        ws.cell(row=hrow, column=2 + j, value=name)
        ws.cell(row=hrow + 1 + j, column=1, value=name).font = Font(name=FONT, bold=True)
    _style_header(ws, hrow, len(names) + 1)

    for i in range(len(names)):
        cli = get_column_letter(2 + i)
        Ri = f"Returns!${cli}$2:${cli}$100000"
        for j in range(len(names)):
            clj = get_column_letter(2 + j)
            Rj = f"Returns!${clj}$2:${clj}$100000"
            r = hrow + 1 + i
            c = 2 + j
            f = f"=IFERROR(CORREL({Ri},{Rj}),\"\")"
            cell = ws.cell(row=r, column=c, value=f)
            cell.number_format = "0.00"
            cell.font = BASE_FONT
            cell.border = BORDER
    for c in range(1, len(names) + 2):
        ws.column_dimensions[get_column_letter(c)].width = 13


def _write_check(ws, prices: pd.DataFrame):
    """검증 시트: 같은 지표를 pandas로 계산한 '값' (공식과 대조용)."""
    rets = prices.pct_change()
    ann_vol = rets.std() * np.sqrt(TRADING_DAYS)
    ann_ret = rets.mean() * TRADING_DAYS
    sharpe = (ann_ret - RISK_FREE) / ann_vol
    latest = prices.iloc[-1]
    r1d = prices.iloc[-1] / prices.iloc[-2] - 1

    ws.cell(row=1, column=1,
            value="🔎 파이썬(pandas) 계산값 — Analysis 시트 공식 결과와 일치해야 함").font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"(rf={RISK_FREE:.0%} 기준, 계산 시각: {datetime.now():%Y-%m-%d %H:%M})").font = Font(name=FONT, italic=True, color="888888", size=9)
    headers = ["자산", "최신가", "1일%", "연환산수익%", "연환산변동성%", "샤프비율"]
    hrow = 4
    for c, h in enumerate(headers, start=1):
        ws.cell(row=hrow, column=c, value=h)
    _style_header(ws, hrow, len(headers))

    for i, name in enumerate(prices.columns):
        r = hrow + 1 + i
        ws.cell(row=r, column=1, value=name).font = Font(name=FONT, bold=True)
        vals = [float(latest[name]), float(r1d[name]),
                float(ann_ret[name]), float(ann_vol[name]), float(sharpe[name])]
        fmts = ["#,##0.00", "0.00%", "0.00%", "0.00%", "0.00"]
        for k, (v, fmt) in enumerate(zip(vals, fmts), start=2):
            cell = ws.cell(row=r, column=k, value=v)
            cell.number_format = fmt
            cell.font = BASE_FONT
    ws.column_dimensions["A"].width = 14
    for c in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 14


def _write_alerts(ws, alerts, causes):
    """자동알림 시트: 파이썬이 판정한 이상변동 + 뉴스 원인분석 (값)."""
    causes = causes or {}
    ws.cell(row=1, column=1, value="🚨 자동 이상변동 알림 (자산별 최신 정규장 종가 기준)").font = TITLE_FONT
    flagged_n = sum(1 for a in alerts if a.get("flagged"))
    ws.cell(row=2, column=1,
            value=f"FLAG {flagged_n}건 · 생성 {datetime.now():%Y-%m-%d %H:%M} · |Z|≥임계 또는 |1일%|≥고정임계 시 경보").font = Font(name=FONT, italic=True, color="888888", size=9)

    headers = ["자산", "기준일", "최신종가", "1일%", "60일σ", "Z", "고정임계", "FLAG", "방향", "원인분석(뉴스)"]
    hrow = 4
    for c, h in enumerate(headers, start=1):
        ws.cell(row=hrow, column=c, value=h)
    _style_header(ws, hrow, len(headers))

    red = PatternFill("solid", fgColor="FEE2E2")
    for i, a in enumerate(alerts):
        r = hrow + 1 + i
        ws.cell(row=r, column=1, value=a["name"]).font = Font(name=FONT, bold=True)
        if a.get("error"):
            ws.cell(row=r, column=10, value=f"수집실패: {a['error']}").font = BASE_FONT
            continue
        vals = [
            (2, a["date"], None), (3, a["last_close"], "#,##0.00"),
            (4, a["ret_1d"], "0.00%"), (5, a["sigma"], "0.00%"),
            (6, a["z"], "0.00"), (7, a["fixed"], "0.0%"),
            (8, "FLAG" if a["flagged"] else "", None),
            (9, a["direction"], None),
            (10, causes.get(a["name"], "") if a["flagged"] else "", None),
        ]
        for c, v, fmt in vals:
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = BASE_FONT
            if fmt:
                cell.number_format = fmt
            if c == 10:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        if a["flagged"]:
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).fill = red

    ws.freeze_panes = "B5"
    widths = [14, 11, 13, 9, 9, 7, 9, 7, 7, 70]
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w


def _write_analyst(ws, analysis_text: str):
    """애널리스트분석 시트: 심층 메모(markdown)를 줄 단위로 표시."""
    ws.cell(row=1, column=1, value="🧠 심층 애널리스트 분석").font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"생성 {datetime.now():%Y-%m-%d %H:%M}").font = Font(name=FONT, italic=True, color="888888", size=9)
    ws.column_dimensions["A"].width = 110
    r = 4
    for line in (analysis_text or "").split("\n"):
        cell = ws.cell(row=r, column=1, value=line)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if line.startswith("## "):
            cell.value = line[3:]
            cell.font = Font(name=FONT, bold=True, size=13, color="4F46E5")
        elif line.strip().startswith(("1.", "2.", "3.", "- ")):
            cell.font = BASE_FONT
        else:
            cell.font = BASE_FONT
        r += 1


def _append_alert_history(wb, alerts, causes):
    """알림기록 시트: FLAG된 자산을 날짜별로 계속 아래에 누적 (덮어쓰지 않음)."""
    causes = causes or {}
    flagged = [a for a in alerts if a.get("flagged")]
    if not flagged:
        return

    if "알림기록" in wb.sheetnames:
        ws = wb["알림기록"]
    else:
        ws = wb.create_sheet("알림기록")
        headers = ["기록일시", "기준일", "자산", "1일%", "Z", "방향", "원인분석(뉴스)"]
        for c, h in enumerate(headers, start=1):
            ws.cell(row=1, column=c, value=h)
        _style_header(ws, 1, len(headers))
        ws.freeze_panes = "A2"
        for c, w in enumerate([17, 11, 14, 9, 7, 7, 95], start=1):
            ws.column_dimensions[get_column_letter(c)].width = w

    # 이미 기록된 (기준일, 자산)은 중복 방지 (같은 날 여러 번 실행해도 안 쌓임)
    existing = set()
    for r in range(2, ws.max_row + 1):
        existing.add((str(ws.cell(r, 2).value), ws.cell(r, 3).value))

    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    for a in flagged:
        if (str(a["date"]), a["name"]) in existing:
            continue
        r = ws.max_row + 1
        ws.cell(row=r, column=1, value=now).font = BASE_FONT
        ws.cell(row=r, column=2, value=a["date"]).font = BASE_FONT
        ws.cell(row=r, column=3, value=a["name"]).font = Font(name=FONT, bold=True)
        c4 = ws.cell(row=r, column=4, value=a["ret_1d"]); c4.number_format = "0.00%"; c4.font = BASE_FONT
        c5 = ws.cell(row=r, column=5, value=a["z"]); c5.number_format = "0.00"; c5.font = BASE_FONT
        ws.cell(row=r, column=6, value=a["direction"]).font = BASE_FONT
        cc = ws.cell(row=r, column=7, value=causes.get(a["name"], ""))
        cc.font = BASE_FONT
        cc.alignment = Alignment(wrap_text=True, vertical="top")


def build_analysis_workbook(tickers, path=None, period="2y", alerts=None, causes=None,
                            analysis_text=None):
    """
    분석 워크북을 생성/갱신.
    alerts: modules.alerts.compute_alerts() 결과 (있으면 자동알림 시트 생성)
    causes: modules.alerts.analyze_causes() 결과
    analysis_text: modules.analyst.generate_deep_analysis() 결과 (있으면 애널리스트분석 시트)
    반환: (파일경로, 에러메시지 or None)
    """
    path = path or DEFAULT_PATH
    try:
        try:
            from config import INCLUDE_TODAY_IN_EXCEL as _inc
        except Exception:
            _inc = False
        prices, failed = _download_prices(tickers, period=period, include_today=_inc)
        if prices.empty:
            return None, "가격 데이터를 하나도 못 받음"

        # 기존 파일 있으면 열어서 관리 시트만 교체 (사용자 메모 시트 등 보존)
        if os.path.exists(path):
            wb = load_workbook(path)
            for sh in MANAGED_SHEETS:
                if sh in wb.sheetnames:
                    del wb[sh]
        else:
            wb = Workbook()
            wb.remove(wb.active)  # 기본 시트 제거

        # 순서대로 생성
        if analysis_text:
            _write_analyst(wb.create_sheet("애널리스트분석"), analysis_text)
        if alerts:
            _write_alerts(wb.create_sheet("자동알림"), alerts, causes)
        _write_prices(wb.create_sheet("Prices"), prices)
        _write_returns(wb.create_sheet("Returns"), prices)
        _write_analysis(wb.create_sheet("Analysis"), prices)
        _write_correlation(wb.create_sheet("Correlation"), prices)
        _write_check(wb.create_sheet("검증(Python)"), prices)

        # 알림 누적 기록 (append 전용, 덮어쓰지 않음)
        if alerts:
            _append_alert_history(wb, alerts, causes)

        # 보기 좋게 가장 중요한 시트를 맨 앞으로
        for cand in ("애널리스트분석", "자동알림", "Analysis"):
            if cand in wb.sheetnames:
                wb.move_sheet(cand, -(wb.sheetnames.index(cand)))
                break

        wb.save(path)
        return path, (f"일부 종목 실패: {', '.join(failed)}" if failed else None)

    except Exception as e:
        return None, f"{type(e).__name__}: {e}"
